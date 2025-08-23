# src/AnalisisSuelosValidados/infrastructure/excel_export_validados_service.py
import pandas as pd
import io
from datetime import datetime
from typing import List, Dict, Any, Optional
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.worksheet.worksheet import Worksheet

class ExcelExportValidadosService:
    """Servicio para exportar datos de análisis de suelos VALIDADOS a archivos Excel"""
    
    def __init__(self):
        self.font_title = Font(name='Arial', size=14, bold=True, color='FFFFFF')
        self.font_header = Font(name='Arial', size=11, bold=True, color='FFFFFF')
        self.font_data = Font(name='Arial', size=10)
        self.font_info = Font(name='Arial', size=11, bold=True)
        
        self.fill_title = PatternFill(start_color='2E7D32', end_color='2E7D32', fill_type='solid')  # Verde para validados
        self.fill_header = PatternFill(start_color='388E3C', end_color='388E3C', fill_type='solid')  # Verde más claro
        self.fill_info = PatternFill(start_color='66BB6A', end_color='66BB6A', fill_type='solid')  # Verde aún más claro
        
        self.border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        self.alignment_center = Alignment(horizontal='center', vertical='center')
        self.alignment_left = Alignment(horizontal='left', vertical='center')

    def exportar_validados_a_excel(
        self, 
        datos: List[Dict[str, Any]], 
        usuario_info: Dict[str, Any], 
        correo_usuario: str
    ) -> Optional[io.BytesIO]:
        """
        Exporta los análisis validados de un usuario a un archivo Excel con formato profesional.
        
        Args:
            datos: Lista de diccionarios con los datos de análisis validados
            usuario_info: Información del usuario
            correo_usuario: Correo del usuario
            
        Returns:
            BytesIO buffer con el archivo Excel o None si hay error
        """
        try:
            if not datos:
                return None
            
            # Crear workbook y worksheet
            wb = Workbook()
            ws = wb.active
            ws.title = "Análisis Validados"
            
            # Configurar el worksheet
            self._configurar_worksheet_inicial(ws, usuario_info, correo_usuario, len(datos))
            
            # Preparar datos para DataFrame
            df_data = self._preparar_datos_para_dataframe(datos)
            
            # Crear DataFrame
            df = pd.DataFrame(df_data)
            
            # Agregar datos al worksheet
            fila_inicio_datos = 8  # Después de la información del usuario
            self._agregar_datos_al_worksheet(ws, df, fila_inicio_datos)
            
            # Aplicar estilos
            self._aplicar_estilos(ws, df, fila_inicio_datos)
            
            # Ajustar anchos de columna
            self._ajustar_anchos_columnas(ws, df)
            
            # Guardar en buffer
            buffer = io.BytesIO()
            wb.save(buffer)
            buffer.seek(0)
            
            return buffer
            
        except Exception as e:
            print(f"Error al crear archivo Excel de validados: {str(e)}")
            return None

    def _configurar_worksheet_inicial(
        self, 
        ws: Worksheet, 
        usuario_info: Dict[str, Any], 
        correo_usuario: str, 
        total_registros: int
    ):
        """Configura la información inicial del worksheet"""
        
        # Título principal
        ws.merge_cells('A1:E1')
        ws['A1'] = "REPORTE DE ANÁLISIS DE SUELOS VALIDADOS"
        ws['A1'].font = self.font_title
        ws['A1'].fill = self.fill_title
        ws['A1'].alignment = self.alignment_center
        
        # Información del usuario
        fecha_actual = datetime.now().strftime("%d/%m/%Y %H:%M")
        
        ws['A3'] = "INFORMACIÓN DEL USUARIO - ANÁLISIS VALIDADOS"
        ws['A3'].font = self.font_info
        ws['A3'].fill = self.fill_info
        
        ws['A4'] = f"Correo: {correo_usuario}"
        ws['A5'] = f"Nombre: {usuario_info.get('nombre', '')} {usuario_info.get('apellido', '')}"
        ws['A6'] = f"Total de análisis validados: {total_registros}"
        ws['A7'] = f"Fecha de generación: {fecha_actual}"

    def _preparar_datos_para_dataframe(self, datos: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
        """Prepara los datos organizándolos en el orden deseado para el DataFrame"""
        
        # Definir el orden y nombres de las columnas para análisis validados
        columnas_ordenadas = [
            ('numero_registro', 'No.'),
            ('id', 'ID'),
            ('municipio_cuadernillo', 'Municipio'),
            ('localidad_cuadernillo', 'Localidad'),
            ('nombre_productor', 'Nombre del Productor'),
            ('tel_productor', 'Teléfono Productor'),
            ('correo_productor', 'Correo Productor'),
            ('nombre_tecnico', 'Nombre del Técnico'),
            ('tel_tecnico', 'Teléfono Técnico'),
            ('correo_tecnico', 'Correo Técnico'),
            ('cultivo_anterior', 'Cultivo Anterior'),
            ('cultivo_establecer', 'Cultivo a Establecer'),
            ('manejo', 'Manejo'),
            ('tipo_vegetacion', 'Tipo de Vegetación'),
            ('parcela', 'Parcela'),
            ('coordenada_x', 'Coordenada X'),
            ('coordenada_y', 'Coordenada Y'),
            ('elevacion_msnm', 'Elevación (msnm)'),
            ('profundidad_muestreo', 'Profundidad de Muestreo'),
            ('fecha_muestreo', 'Fecha de Muestreo'),
            ('muestra', 'Muestra'),
            ('reemplazo', 'Reemplazo'),
            ('ddr', 'DDR'),
            ('cader', 'CADER'),
            ('clave', 'Clave'),
            ('numero', 'Número'),
            ('clave_estatal', 'Clave Estatal'),
            ('clave_municipio', 'Clave Municipio'),
            ('clave_munip', 'Clave Munip'),
            ('clave_localidad', 'Clave Localidad'),
            ('estado_cuadernillo', 'Estado Cuadernillo'),
            ('recuento_curp_renapo', 'Recuento CURP RENAPO'),
            ('extraccion_edo', 'Extracción Edo'),
            ('nombre_revisor', 'Nombre del Revisor'),
            ('fecha_validacion', 'Fecha de Validación'),
            ('fecha_creacion', 'Fecha de Creación')
        ]
        
        datos_procesados = []
        for dato in datos:
            fila_procesada = {}
            for campo_original, nombre_columna in columnas_ordenadas:
                valor = dato.get(campo_original, '')
                # Convertir None a string vacío y asegurar que todos los valores sean strings
                fila_procesada[nombre_columna] = str(valor) if valor is not None else ''
            
            datos_procesados.append(fila_procesada)
        
        return datos_procesados

    def _agregar_datos_al_worksheet(self, ws: Worksheet, df: pd.DataFrame, fila_inicio: int):
        """Agrega los datos del DataFrame al worksheet"""
        
        # Agregar encabezados
        for col_num, column_name in enumerate(df.columns, 1):
            cell = ws.cell(row=fila_inicio, column=col_num, value=column_name)
            cell.font = self.font_header
            cell.fill = self.fill_header
            cell.alignment = self.alignment_center
            cell.border = self.border
        
        # Agregar datos
        for row_num, row_data in enumerate(df.itertuples(index=False), fila_inicio + 1):
            for col_num, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_num, column=col_num, value=value)
                cell.font = self.font_data
                cell.alignment = self.alignment_left
                cell.border = self.border
                
                # Alternar colores de fila para mejor legibilidad
                if row_num % 2 == 0:
                    cell.fill = PatternFill(start_color='F1F8E9', end_color='F1F8E9', fill_type='solid')

    def _aplicar_estilos(self, ws: Worksheet, df: pd.DataFrame, fila_inicio: int):
        """Aplica estilos adicionales al worksheet"""
        
        # Aplicar bordes a la información del usuario (filas 1-7)
        for row in range(1, fila_inicio):
            for col in range(1, 6):
                cell = ws.cell(row=row, column=col)
                if cell.value is not None and str(cell.value).strip():
                    cell.border = self.border
                    cell.font = self.font_data

    def _ajustar_anchos_columnas(self, ws: Worksheet, df: pd.DataFrame):
        """Ajusta automáticamente los anchos de las columnas"""
        
        for col_idx, column_name in enumerate(df.columns, 1):
            max_length = len(str(column_name))
            
            for row_idx in range(1, ws.max_row + 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                try:
                    if cell.value is not None:
                        cell_length = len(str(cell.value))
                        if cell_length > max_length:
                            max_length = cell_length
                except:
                    pass
            
            adjusted_width = min(max(max_length + 2, 12), 50)
            
            from openpyxl.utils import get_column_letter
            column_letter = get_column_letter(col_idx)
            ws.column_dimensions[column_letter].width = adjusted_width

    def crear_excel_simple(self, datos: List[Dict[str, Any]]) -> Optional[io.BytesIO]:
        """
        Método alternativo para crear un Excel simple usando solo pandas
        """
        try:
            if not datos:
                return None
            
            df_data = self._preparar_datos_para_dataframe(datos)
            df = pd.DataFrame(df_data)
            
            buffer = io.BytesIO()
            
            with pd.ExcelWriter(buffer, engine='openpyxl') as writer:
                df.to_excel(writer, sheet_name='Análisis Validados', index=False)
            
            buffer.seek(0)
            return buffer
            
        except Exception as e:
            print(f"Error al crear Excel simple de validados: {str(e)}")
            return None

    # MÉTODOS PARA TODOS LOS VALIDADOS
    
    def exportar_todos_validados_a_excel(
        self, 
        datos: List[Dict[str, Any]], 
        total_usuarios: int,
        usuarios_con_validados: List[str]
    ) -> Optional[io.BytesIO]:
        """
        Exporta TODOS los análisis validados de TODOS los usuarios a un archivo Excel.
        
        Args:
            datos: Lista de diccionarios con los datos de análisis validados de todos los usuarios
            total_usuarios: Número total de usuarios con análisis validados
            usuarios_con_validados: Lista de correos de usuarios con validados
            
        Returns:
            BytesIO buffer con el archivo Excel o None si hay error
        """
        try:
            if not datos:
                return None
            
            wb = Workbook()
            ws = wb.active
            ws.title = "Todos los Análisis Validados"
            
            self._configurar_worksheet_todos_validados(ws, total_usuarios, usuarios_con_validados, len(datos))
            
            df_data = self._preparar_datos_todos_validados_para_dataframe(datos)
            df = pd.DataFrame(df_data)
            
            fila_inicio_datos = 10
            self._agregar_datos_al_worksheet(ws, df, fila_inicio_datos)
            self._aplicar_estilos(ws, df, fila_inicio_datos)
            self._ajustar_anchos_columnas(ws, df)
            
            buffer = io.BytesIO()
            wb.save(buffer)
            buffer.seek(0)
            
            return buffer
            
        except Exception as e:
            print(f"Error al crear archivo Excel de todos los validados: {str(e)}")
            return None

    def _configurar_worksheet_todos_validados(
        self, 
        ws: Worksheet, 
        total_usuarios: int,
        usuarios_con_validados: List[str],
        total_registros: int
    ):
        """Configura la información inicial del worksheet para todos los validados"""
        
        ws.merge_cells('A1:F1')
        ws['A1'] = "REPORTE COMPLETO DE ANÁLISIS DE SUELOS VALIDADOS - TODOS LOS USUARIOS"
        ws['A1'].font = self.font_title
        ws['A1'].fill = self.fill_title
        ws['A1'].alignment = self.alignment_center
        
        fecha_actual = datetime.now().strftime("%d/%m/%Y %H:%M")
        
        ws['A3'] = "INFORMACIÓN GENERAL DEL REPORTE - ANÁLISIS VALIDADOS"
        ws['A3'].font = self.font_info
        ws['A3'].fill = self.fill_info
        
        ws['A4'] = f"Total de usuarios con análisis validados: {total_usuarios}"
        ws['A5'] = f"Total de análisis validados en el sistema: {total_registros}"
        ws['A6'] = f"Fecha de generación del reporte: {fecha_actual}"
        ws['A7'] = f"Generado por: Sistema de Análisis de Suelos"
        
        ws['A9'] = "USUARIOS CON ANÁLISIS VALIDADOS (Primeros 10):"
        ws['A9'].font = Font(name='Arial', size=10, bold=True)
        
        usuarios_mostrar = usuarios_con_validados[:10] if len(usuarios_con_validados) > 10 else usuarios_con_validados
        usuarios_texto = ", ".join(usuarios_mostrar)
        if len(usuarios_con_validados) > 10:
            usuarios_texto += f" ... y {len(usuarios_con_validados) - 10} más"
        
        if len(usuarios_texto) > 100:
            ws['B9'] = usuarios_texto[:100] + "..."
        else:
            ws['B9'] = usuarios_texto

    def _preparar_datos_todos_validados_para_dataframe(self, datos: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
        """Prepara los datos organizándolos para el DataFrame de todos los validados"""
        
        columnas_ordenadas = [
            ('numero_registro', 'No.'),
            ('id', 'ID Análisis'),
            ('usuario_correo', 'Correo del Usuario'),
            ('usuario_nombre_completo', 'Nombre del Usuario'),
            ('usuario_rol', 'Rol Usuario'),
            ('municipio_cuadernillo', 'Municipio'),
            ('localidad_cuadernillo', 'Localidad'),
            ('nombre_productor', 'Nombre del Productor'),
            ('tel_productor', 'Teléfono Productor'),
            ('correo_productor', 'Correo Productor'),
            ('nombre_tecnico', 'Nombre del Técnico'),
            ('tel_tecnico', 'Teléfono Técnico'),
            ('correo_tecnico', 'Correo Técnico'),
            ('cultivo_anterior', 'Cultivo Anterior'),
            ('cultivo_establecer', 'Cultivo a Establecer'),
            ('manejo', 'Manejo'),
            ('tipo_vegetacion', 'Tipo de Vegetación'),
            ('parcela', 'Parcela'),
            ('coordenada_x', 'Coordenada X'),
            ('coordenada_y', 'Coordenada Y'),
            ('elevacion_msnm', 'Elevación (msnm)'),
            ('profundidad_muestreo', 'Profundidad de Muestreo'),
            ('fecha_muestreo', 'Fecha de Muestreo'),
            ('muestra', 'Muestra'),
            ('reemplazo', 'Reemplazo'),
            ('ddr', 'DDR'),
            ('cader', 'CADER'),
            ('clave', 'Clave'),
            ('numero', 'Número'),
            ('clave_estatal', 'Clave Estatal'),
            ('clave_municipio', 'Clave Municipio'),
            ('clave_munip', 'Clave Munip'),
            ('clave_localidad', 'Clave Localidad'),
            ('estado_cuadernillo', 'Estado Cuadernillo'),
            ('recuento_curp_renapo', 'Recuento CURP RENAPO'),
            ('extraccion_edo', 'Extracción Edo'),
            ('nombre_revisor', 'Nombre del Revisor'),
            ('fecha_validacion', 'Fecha de Validación'),
            ('fecha_creacion', 'Fecha de Creación')
        ]
        
        datos_procesados = []
        for dato in datos:
            fila_procesada = {}
            for campo_original, nombre_columna in columnas_ordenadas:
                valor = dato.get(campo_original, '')
                fila_procesada[nombre_columna] = str(valor) if valor is not None else ''
            
            datos_procesados.append(fila_procesada)
        
        return datos_procesados

    def crear_excel_simple_todos(self, datos: List[Dict[str, Any]]) -> Optional[io.BytesIO]:
        """
        Método alternativo para crear un Excel simple de todos los validados
        """
        try:
            if not datos:
                return None
            
            df_data = self._preparar_datos_todos_validados_para_dataframe(datos)
            df = pd.DataFrame(df_data)
            
            buffer = io.BytesIO()
            
            with pd.ExcelWriter(buffer, engine='openpyxl') as writer:
                df.to_excel(writer, sheet_name='Todos los Análisis Validados', index=False)
            
            buffer.seek(0)
            return buffer
            
        except Exception as e:
            print(f"Error al crear Excel simple de todos los validados: {str(e)}")
            return None