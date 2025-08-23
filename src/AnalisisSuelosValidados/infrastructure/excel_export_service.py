# src/AnalisisSuelosValidados/infrastructure/excel_export_service.py
import pandas as pd
import io
from datetime import datetime
from typing import List, Dict, Any, Optional
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.worksheet.worksheet import Worksheet

class ExcelExportService:
    """Servicio para exportar datos de análisis de suelos a archivos Excel"""
    
    def __init__(self):
        self.font_title = Font(name='Arial', size=14, bold=True, color='FFFFFF')
        self.font_header = Font(name='Arial', size=11, bold=True, color='FFFFFF')
        self.font_data = Font(name='Arial', size=10)
        self.font_info = Font(name='Arial', size=11, bold=True)
        
        self.fill_title = PatternFill(start_color='2E86AB', end_color='2E86AB', fill_type='solid')
        self.fill_header = PatternFill(start_color='A23B72', end_color='A23B72', fill_type='solid')
        self.fill_info = PatternFill(start_color='F18F01', end_color='F18F01', fill_type='solid')
        
        self.border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        self.alignment_center = Alignment(horizontal='center', vertical='center')
        self.alignment_left = Alignment(horizontal='left', vertical='center')

    def exportar_pendientes_a_excel(
        self, 
        datos: List[Dict[str, Any]], 
        usuario_info: Dict[str, Any], 
        correo_usuario: str
    ) -> Optional[io.BytesIO]:
        """
        Exporta los análisis pendientes a un archivo Excel con formato profesional.
        
        Args:
            datos: Lista de diccionarios con los datos de análisis
            usuario_info: Información del usuario
            correo_usuario: Correo del usuario
            
        Returns:
            BytesIO buffer con el archivo Excel o None si hay error
        """
        try:
            if not datos:
                return None
            
            # Crear workbook y worksheet
            wb = Workbook()
            ws = wb.active
            ws.title = "Análisis Pendientes"
            
            # Configurar el worksheet
            self._configurar_worksheet_inicial(ws, usuario_info, correo_usuario, len(datos))
            
            # Preparar datos para DataFrame
            df_data = self._preparar_datos_para_dataframe(datos)
            
            # Crear DataFrame
            df = pd.DataFrame(df_data)
            
            # Agregar datos al worksheet
            fila_inicio_datos = 8  # Después de la información del usuario
            self._agregar_datos_al_worksheet(ws, df, fila_inicio_datos)
            
            # Aplicar estilos
            self._aplicar_estilos(ws, df, fila_inicio_datos)
            
            # Ajustar anchos de columna
            self._ajustar_anchos_columnas(ws, df)
            
            # Guardar en buffer
            buffer = io.BytesIO()
            wb.save(buffer)
            buffer.seek(0)
            
            return buffer
            
        except Exception as e:
            print(f"Error al crear archivo Excel: {str(e)}")
            return None

    def _configurar_worksheet_inicial(
        self, 
        ws: Worksheet, 
        usuario_info: Dict[str, Any], 
        correo_usuario: str, 
        total_registros: int
    ):
        """Configura la información inicial del worksheet"""
        
        # Título principal - usar un rango más pequeño para evitar conflictos
        ws.merge_cells('A1:E1')
        ws['A1'] = "REPORTE DE ANÁLISIS DE SUELOS PENDIENTES"
        ws['A1'].font = self.font_title
        ws['A1'].fill = self.fill_title
        ws['A1'].alignment = self.alignment_center
        
        # Información del usuario
        fecha_actual = datetime.now().strftime("%d/%m/%Y %H:%M")
        
        ws['A3'] = "INFORMACIÓN DEL USUARIO QUIN REGISTRO ESTE EXCEL CON STATUS PENDIENTE PARA QUE ESTE EN REVISION"
        ws['A3'].font = self.font_info
        ws['A3'].fill = self.fill_info
        
        ws['A4'] = f"Correo: {correo_usuario}"
        ws['A5'] = f"Nombre: {usuario_info.get('nombre', '')} {usuario_info.get('apellido', '')}"
        ws['A6'] = f"Total de análisis pendientes: {total_registros}"
        ws['A7'] = f"Fecha de generación: {fecha_actual}"

    def _preparar_datos_para_dataframe(self, datos: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
        """Prepara los datos organizándolos en el orden deseado para el DataFrame"""
        
        # Definir el orden y nombres de las columnas
        columnas_ordenadas = [
            ('numero_registro', 'No.'),
            ('id', 'ID'),
            ('municipio_cuadernillo', 'Municipio'),
            ('localidad_cuadernillo', 'Localidad'),
            ('nombre_productor', 'Nombre del Productor'),
            ('tel_productor', 'Teléfono Productor'),
            ('correo_productor', 'Correo Productor'),
            ('nombre_tecnico', 'Nombre del Técnico'),
            ('tel_tecnico', 'Teléfono Técnico'),
            ('correo_tecnico', 'Correo Técnico'),
            ('cultivo_anterior', 'Cultivo Anterior'),
            ('cultivo_establecer', 'Cultivo a Establecer'),
            ('manejo', 'Manejo'),
            ('tipo_vegetacion', 'Tipo de Vegetación'),
            ('parcela', 'Parcela'),
            ('coordenada_x', 'Coordenada X'),
            ('coordenada_y', 'Coordenada Y'),
            ('elevacion_msnm', 'Elevación (msnm)'),
            ('profundidad_muestreo', 'Profundidad de Muestreo'),
            ('fecha_muestreo', 'Fecha de Muestreo'),
            ('muestra', 'Muestra'),
            ('reemplazo', 'Reemplazo'),
            ('ddr', 'DDR'),
            ('cader', 'CADER'),
            ('clave', 'Clave'),
            ('numero', 'Número'),
            ('clave_estatal', 'Clave Estatal'),
            ('clave_municipio', 'Clave Municipio'),
            ('clave_munip', 'Clave Munip'),
            ('clave_localidad', 'Clave Localidad'),
            ('estado_cuadernillo', 'Estado Cuadernillo'),
            ('recuento_curp_renapo', 'Recuento CURP RENAPO'),
            ('extraccion_edo', 'Extracción Edo'),
            ('nombre_revisor', 'Nombre del Revisor'),
            ('fecha_creacion', 'Fecha de Creación')
        ]
        
        datos_procesados = []
        for dato in datos:
            fila_procesada = {}
            for campo_original, nombre_columna in columnas_ordenadas:
                valor = dato.get(campo_original, '')
                # Convertir None a string vacío y asegurar que todos los valores sean strings
                fila_procesada[nombre_columna] = str(valor) if valor is not None else ''
            
            datos_procesados.append(fila_procesada)
        
        return datos_procesados

    def _agregar_datos_al_worksheet(self, ws: Worksheet, df: pd.DataFrame, fila_inicio: int):
        """Agrega los datos del DataFrame al worksheet"""
        
        # Agregar encabezados
        for col_num, column_name in enumerate(df.columns, 1):
            cell = ws.cell(row=fila_inicio, column=col_num, value=column_name)
            cell.font = self.font_header
            cell.fill = self.fill_header
            cell.alignment = self.alignment_center
            cell.border = self.border
        
        # Agregar datos
        for row_num, row_data in enumerate(df.itertuples(index=False), fila_inicio + 1):
            for col_num, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_num, column=col_num, value=value)
                cell.font = self.font_data
                cell.alignment = self.alignment_left
                cell.border = self.border
                
                # Alternar colores de fila para mejor legibilidad
                if row_num % 2 == 0:
                    cell.fill = PatternFill(start_color='F8F9FA', end_color='F8F9FA', fill_type='solid')

    def _aplicar_estilos(self, ws: Worksheet, df: pd.DataFrame, fila_inicio: int):
        """Aplica estilos adicionales al worksheet"""
        
        # Aplicar bordes a la información del usuario (filas 1-7)
        for row in range(1, fila_inicio):
            for col in range(1, 6):  # Solo las primeras 5 columnas para evitar conflictos
                cell = ws.cell(row=row, column=col)
                if cell.value is not None and str(cell.value).strip():  # Solo aplicar si hay contenido real
                    cell.border = self.border
                    cell.font = self.font_data

    def _ajustar_anchos_columnas(self, ws: Worksheet, df: pd.DataFrame):
        """Ajusta automáticamente los anchos de las columnas"""
        
        # Obtener el rango de columnas basado en el DataFrame
        for col_idx, column_name in enumerate(df.columns, 1):
            max_length = len(str(column_name))  # Empezar con la longitud del encabezado
            
            # Revisar todas las celdas de la columna para encontrar la más larga
            for row_idx in range(1, ws.max_row + 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                try:
                    if cell.value is not None:
                        cell_length = len(str(cell.value))
                        if cell_length > max_length:
                            max_length = cell_length
                except:
                    pass
            
            # Establecer un ancho mínimo y máximo
            adjusted_width = min(max(max_length + 2, 12), 50)
            
            # Obtener la letra de la columna
            from openpyxl.utils import get_column_letter
            column_letter = get_column_letter(col_idx)
            ws.column_dimensions[column_letter].width = adjusted_width

    def crear_excel_simple(self, datos: List[Dict[str, Any]]) -> Optional[io.BytesIO]:
        """
        Método alternativo para crear un Excel simple usando solo pandas
        en caso de que la versión completa presente problemas
        """
        try:
            if not datos:
                return None
            
            # Preparar datos
            df_data = self._preparar_datos_para_dataframe(datos)
            df = pd.DataFrame(df_data)
            
            # Crear buffer
            buffer = io.BytesIO()
            
            # Crear archivo Excel simple
            with pd.ExcelWriter(buffer, engine='openpyxl') as writer:
                df.to_excel(writer, sheet_name='Análisis Pendientes', index=False)
            
            buffer.seek(0)
            return buffer
            
        except Exception as e:
            print(f"Error al crear Excel simple: {str(e)}")
            return None