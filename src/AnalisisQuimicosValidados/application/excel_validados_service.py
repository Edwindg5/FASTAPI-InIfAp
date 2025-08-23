# src/AnalisisQuimicosValidados/application/excel_validados_service.py
import pandas as pd
from io import BytesIO
from typing import Optional, List, Dict, Any
from sqlalchemy.orm import Session
from datetime import datetime
import traceback

from src.AnalisisQuimicosValidados.application.todos_validados_service import (
    obtener_todos_los_validados,
    obtener_validados_por_usuario_completo,
    buscar_validados_por_filtros
)
from src.Users.infrastructure.users_model import Users


def generar_excel_todos_validados(db: Session) -> Optional[BytesIO]:
    """
    Genera un archivo Excel con TODOS los análisis químicos validados.
    
    Args:
        db (Session): Sesión de base de datos
        
    Returns:
        Optional[BytesIO]: Archivo Excel en memoria o None si hay error
    """
    try:
        print("=== GENERANDO EXCEL TODOS LOS VALIDADOS ===")
        
        # Obtener todos los análisis validados (sin límite)
        resultado = obtener_todos_los_validados(db, limit=None)
        
        if not resultado["success"] or not resultado["analisis_validados"]:
            print("No se encontraron análisis validados para exportar")
            return None
        
        analisis_validados = resultado["analisis_validados"]
        estadisticas = resultado["estadisticas"]
        
        print(f"Análisis validados a exportar: {len(analisis_validados)}")
        
        # Convertir los datos a formato Excel
        datos_excel = []
        for analisis in analisis_validados:
            try:
                # Información del usuario validador
                usuario_validador = analisis.get("usuario_validador", {})
                correo_validador = usuario_validador.get("correo", "Sin especificar") if usuario_validador else "Sin especificar"
                nombre_validador = usuario_validador.get("nombre", "Sin especificar") if usuario_validador else "Sin especificar"
                
                fila = {
                    # Información básica
                    "ID": analisis["id"],
                    "Usuario Validador ID": analisis["user_id_FK"],
                    "Correo Validador": correo_validador,
                    "Nombre Validador": nombre_validador,
                    
                    # Ubicación y productor
                    "Municipio": analisis["municipio"],
                    "Localidad": analisis["localidad"],
                    "Nombre Productor": analisis["nombre_productor"],
                    "Cultivo Anterior": analisis["cultivo_anterior"],
                    
                    # Propiedades físicas del suelo
                    "Arcilla (%)": analisis["arcilla"],
                    "Limo (%)": analisis["limo"],
                    "Arena (%)": analisis["arena"],
                    "Textura": analisis["textura"],
                    "Densidad Aparente": analisis["da"],
                    
                    # Propiedades químicas básicas
                    "pH": analisis["ph"],
                    "Materia Orgánica (%)": analisis["mo"],
                    "Fósforo (ppm)": analisis["fosforo"],
                    "N Inorgánico (ppm)": analisis["n_inorganico"],
                    
                    # Macronutrientes intercambiables
                    "Potasio (ppm)": analisis["k"],
                    "Magnesio (ppm)": analisis["mg"],
                    "Calcio (ppm)": analisis["ca"],
                    "Sodio (ppm)": analisis["na"],
                    
                    # CIC y acidez
                    "Aluminio (ppm)": analisis["al"],
                    "CIC (meq/100g)": analisis["cic"],
                    "CIC Calculada": analisis["cic_calculada"],
                    "Hidrógeno (ppm)": analisis["h"],
                    
                    # Micronutrientes
                    "Azufre (ppm)": analisis["azufre"],
                    "Hierro (ppm)": analisis["hierro"],
                    "Cobre (ppm)": analisis["cobre"],
                    "Zinc (ppm)": analisis["zinc"],
                    "Manganeso (ppm)": analisis["manganeso"],
                    "Boro (ppm)": analisis["boro"],
                    
                    # Relaciones catiónicas
                    "Ca/Mg": analisis["ca_mg"],
                    "Mg/K": analisis["mg_k"],
                    "Ca/K": analisis["ca_k"],
                    "Ca+Mg/K": analisis["ca_mg_k"],
                    "K/Mg": analisis["k_mg"],
                    
                    # Columnas adicionales
                    "Columna 1": analisis["columna1"],
                    "Columna 2": analisis["columna2"],
                    
                    # Fechas
                    "Fecha Validación": analisis["fecha_validacion"],
                    "Fecha Creación Original": analisis["fecha_creacion"],
                }
                
                datos_excel.append(fila)
                
            except Exception as e:
                print(f"Error procesando análisis {analisis.get('id', 'unknown')}: {e}")
                continue
        
        # Crear DataFrame
        df = pd.DataFrame(datos_excel)
        
        if df.empty:
            print("No se pudieron procesar los datos para Excel")
            return None
        
        # Crear archivo Excel en memoria con múltiples hojas
        buffer = BytesIO()
        
        with pd.ExcelWriter(buffer, engine='openpyxl') as writer:
            
            # HOJA 1: Datos principales
            df.to_excel(
                writer, 
                sheet_name='Análisis Validados', 
                index=False,
                startrow=6  # Dejar espacio para información general
            )
            
            worksheet_main = writer.sheets['Análisis Validados']
            
            # Información general en la hoja principal
            fecha_actual = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            
            worksheet_main['A1'] = "ANÁLISIS QUÍMICOS VALIDADOS - REPORTE COMPLETO"
            worksheet_main['A2'] = f"Fecha de exportación: {fecha_actual}"
            worksheet_main['A3'] = f"Total de análisis validados: {len(datos_excel)}"
            worksheet_main['A4'] = f"Total de usuarios validadores: {len(resultado['usuarios_validadores'])}"
            worksheet_main['A5'] = f"Sistema: Análisis Químicos de Suelos"
            
            # HOJA 2: Estadísticas
            if estadisticas:
                estadisticas_data = []
                
                # Estadísticas generales
                estadisticas_data.append(["ESTADÍSTICAS GENERALES", ""])
                estadisticas_data.append(["Total de análisis", estadisticas.get("total_analisis", 0)])
                estadisticas_data.append(["", ""])
                
                # Top municipios
                if "municipios_top_5" in estadisticas:
                    estadisticas_data.append(["TOP 5 MUNICIPIOS", ""])
                    for municipio, cantidad in estadisticas["municipios_top_5"].items():
                        estadisticas_data.append([municipio, cantidad])
                    estadisticas_data.append(["", ""])
                
                # Usuarios validadores
                if "usuarios_validadores" in estadisticas:
                    estadisticas_data.append(["USUARIOS VALIDADORES", ""])
                    for usuario, cantidad in estadisticas["usuarios_validadores"].items():
                        estadisticas_data.append([usuario, cantidad])
                    estadisticas_data.append(["", ""])
                
                # Estadísticas numéricas
                if "estadisticas_numericas" in estadisticas:
                    estadisticas_data.append(["PROMEDIOS QUÍMICOS", ""])
                    est_num = estadisticas["estadisticas_numericas"]
                    
                    if "ph" in est_num:
                        ph_stats = est_num["ph"]
                        estadisticas_data.append([f"pH - Promedio", f"{ph_stats['promedio']:.2f}"])
                        estadisticas_data.append([f"pH - Rango", f"{ph_stats['minimo']:.2f} - {ph_stats['maximo']:.2f}"])
                    
                    if "materia_organica" in est_num:
                        mo_stats = est_num["materia_organica"]
                        estadisticas_data.append([f"M.O. (%) - Promedio", f"{mo_stats['promedio']:.2f}"])
                        estadisticas_data.append([f"M.O. (%) - Rango", f"{mo_stats['minimo']:.2f} - {mo_stats['maximo']:.2f}"])
                    
                    if "fosforo" in est_num:
                        p_stats = est_num["fosforo"]
                        estadisticas_data.append([f"Fósforo (ppm) - Promedio", f"{p_stats['promedio']:.2f}"])
                        estadisticas_data.append([f"Fósforo (ppm) - Rango", f"{p_stats['minimo']:.2f} - {p_stats['maximo']:.2f}"])
                
                # Crear DataFrame de estadísticas
                df_stats = pd.DataFrame(estadisticas_data, columns=["Concepto", "Valor"])
                df_stats.to_excel(writer, sheet_name='Estadísticas', index=False)
            
            # HOJA 3: Resumen por usuario validador
            if resultado["usuarios_validadores"]:
                resumen_usuarios = []
                for correo_usuario in resultado["usuarios_validadores"]:
                    # Contar análisis por usuario
                    count = sum(1 for a in analisis_validados 
                              if a.get("usuario_validador") and 
                                 a["usuario_validador"].get("correo") == correo_usuario)
                    
                    # Obtener nombre del usuario
                    usuario_info = next(
                        (a["usuario_validador"] for a in analisis_validados 
                         if a.get("usuario_validador") and 
                            a["usuario_validador"].get("correo") == correo_usuario), 
                        {}
                    )
                    
                    resumen_usuarios.append({
                        "Correo Usuario": correo_usuario,
                        "Nombre Usuario": usuario_info.get("nombre", "Sin especificar"),
                        "ID Usuario": usuario_info.get("id", "Sin especificar"),
                        "Total Validaciones": count
                    })
                
                df_usuarios = pd.DataFrame(resumen_usuarios)
                df_usuarios = df_usuarios.sort_values("Total Validaciones", ascending=False)
                df_usuarios.to_excel(writer, sheet_name='Resumen por Usuario', index=False)
            
            # Aplicar formato a las hojas
            aplicar_formato_excel(writer, worksheet_main)
        
        buffer.seek(0)
        print(f"✅ Excel generado exitosamente con {len(datos_excel)} análisis validados")
        
        return buffer
        
    except Exception as e:
        print(f"❌ Error generando Excel de validados: {e}")
        print(f"Traceback: {traceback.format_exc()}")
        return None


def generar_excel_validados_por_usuario(
    user_id_FK: int, 
    db: Session
) -> Optional[BytesIO]:
    """
    Genera un archivo Excel con todos los análisis validados por un usuario específico.
    
    Args:
        user_id_FK (int): ID del usuario validador
        db (Session): Sesión de base de datos
        
    Returns:
        Optional[BytesIO]: Archivo Excel en memoria o None si hay error
    """
    try:
        print(f"=== GENERANDO EXCEL VALIDADOS POR USUARIO {user_id_FK} ===")
        
        # Obtener análisis validados del usuario
        resultado = obtener_validados_por_usuario_completo(user_id_FK, db)
        
        if not resultado["success"] or not resultado["analisis_validados"]:
            print(f"No se encontraron análisis validados para el usuario {user_id_FK}")
            return None
        
        analisis_validados = resultado["analisis_validados"]
        usuario_info = resultado["usuario"]
        
        print(f"Análisis validados a exportar: {len(analisis_validados)}")
        
        # Convertir datos a formato Excel (similar al anterior pero más simple)
        datos_excel = []
        for analisis in analisis_validados:
            fila = {
                "ID": analisis["id"],
                "Municipio": analisis["municipio"],
                "Localidad": analisis["localidad"],
                "Nombre Productor": analisis["nombre_productor"],
                "pH": analisis["ph"],
                "Materia Orgánica (%)": analisis["mo"],
                "Fósforo (ppm)": analisis["fosforo"],
                "Fecha Validación": analisis["fecha_validacion"],
                # Agregar más campos según necesidad
            }
            datos_excel.append(fila)
        
        # Crear DataFrame y Excel
        df = pd.DataFrame(datos_excel)
        buffer = BytesIO()
        
        with pd.ExcelWriter(buffer, engine='openpyxl') as writer:
            df.to_excel(
                writer, 
                sheet_name='Validados por Usuario', 
                index=False,
                startrow=5
            )
            
            worksheet = writer.sheets['Validados por Usuario']
            
            # Información del usuario
            fecha_actual = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            worksheet['A1'] = f"ANÁLISIS VALIDADOS POR USUARIO"
            worksheet['A2'] = f"Usuario: {usuario_info['correo']}"
            worksheet['A3'] = f"Nombre: {usuario_info['nombre']}"
            worksheet['A4'] = f"Total validados: {len(datos_excel)}"
            worksheet['A5'] = f"Fecha de exportación: {fecha_actual}"
            
            aplicar_formato_excel(writer, worksheet)
        
        buffer.seek(0)
        return buffer
        
    except Exception as e:
        print(f"❌ Error generando Excel por usuario: {e}")
        return None


def generar_excel_validados_filtrado(
    db: Session,
    municipio: Optional[str] = None,
    fecha_desde: Optional[datetime] = None,
    fecha_hasta: Optional[datetime] = None,
    usuario_validador_id: Optional[int] = None
) -> Optional[BytesIO]:
    """
    Genera un archivo Excel con análisis validados aplicando filtros.
    
    Args:
        db (Session): Sesión de base de datos
        municipio (Optional[str]): Filtro por municipio
        fecha_desde (Optional[datetime]): Fecha inicio
        fecha_hasta (Optional[datetime]): Fecha fin
        usuario_validador_id (Optional[int]): ID del usuario validador
        
    Returns:
        Optional[BytesIO]: Archivo Excel en memoria o None si hay error
    """
    try:
        print("=== GENERANDO EXCEL VALIDADOS CON FILTROS ===")
        
        # Obtener análisis filtrados (sin límite para Excel)
        resultado = buscar_validados_por_filtros(
            db, municipio, fecha_desde, fecha_hasta, 
            usuario_validador_id, limit=10000, offset=0
        )
        
        if not resultado["success"] or not resultado["analisis_filtrados"]:
            print("No se encontraron análisis con los filtros aplicados")
            return None
        
        analisis_filtrados = resultado["analisis_filtrados"]
        filtros_aplicados = resultado["filtros_aplicados"]
        
        print(f"Análisis filtrados a exportar: {len(analisis_filtrados)}")
        
        # Convertir a Excel
        datos_excel = []
        for analisis in analisis_filtrados:
            fila = {
                "ID": analisis["id"],
                "Municipio": analisis["municipio"],
                "Localidad": analisis["localidad"],
                "Nombre Productor": analisis["nombre_productor"],
                "Fecha Validación": analisis["fecha_validacion"],
                # Agregar más campos según necesidad
            }
            datos_excel.append(fila)
        
        # Crear Excel
        df = pd.DataFrame(datos_excel)
        buffer = BytesIO()
        
        with pd.ExcelWriter(buffer, engine='openpyxl') as writer:
            df.to_excel(
                writer, 
                sheet_name='Validados Filtrados', 
                index=False,
                startrow=8
            )
            
            worksheet = writer.sheets['Validados Filtrados']
            
            # Información de filtros
            fecha_actual = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            worksheet['A1'] = "ANÁLISIS VALIDADOS - REPORTE FILTRADO"
            worksheet['A2'] = f"Fecha de exportación: {fecha_actual}"
            worksheet['A3'] = f"Total encontrados: {len(datos_excel)}"
            worksheet['A4'] = f"Filtro Municipio: {filtros_aplicados['municipio'] or 'Todos'}"
            worksheet['A5'] = f"Filtro Fecha Desde: {filtros_aplicados['fecha_desde'] or 'Sin filtro'}"
            worksheet['A6'] = f"Filtro Fecha Hasta: {filtros_aplicados['fecha_hasta'] or 'Sin filtro'}"
            worksheet['A7'] = f"Filtro Usuario ID: {filtros_aplicados['usuario_validador_id'] or 'Todos'}"
            
            aplicar_formato_excel(writer, worksheet)
        
        buffer.seek(0)
        return buffer
        
    except Exception as e:
        print(f"❌ Error generando Excel filtrado: {e}")
        return None


def aplicar_formato_excel(writer, worksheet):
    """
    Aplica formato básico a las hojas de Excel.
    
    Args:
        writer: ExcelWriter object
        worksheet: Worksheet object
    """
    try:
        from openpyxl.styles import Font, PatternFill, Alignment
        
        # Estilo para títulos
        title_font = Font(bold=True, size=14, color="FFFFFF")
        title_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        
        # Estilo para encabezados
        header_font = Font(bold=True, size=12)
        header_fill = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")
        
        # Aplicar estilo a la primera fila (título principal)
        for cell in worksheet[1]:
            if cell.value:
                cell.font = title_font
                cell.fill = title_fill
                cell.alignment = Alignment(horizontal="center")
        
        # Ajustar ancho de columnas automáticamente
        for column in worksheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            
            adjusted_width = min(max_length + 2, 30)  # Máximo 30 caracteres
            worksheet.column_dimensions[column_letter].width = adjusted_width
        
    except Exception as e:
        print(f"Error aplicando formato Excel: {e}")


def obtener_nombre_archivo_validados(
    tipo: str = "completo",
    usuario_correo: Optional[str] = None,
    filtros: Optional[Dict] = None
) -> str:
    """
    Genera nombre de archivo Excel para diferentes tipos de exportación.
    
    Args:
        tipo (str): Tipo de reporte ('completo', 'usuario', 'filtrado')
        usuario_correo (Optional[str]): Correo del usuario (para tipo 'usuario')
        filtros (Optional[Dict]): Filtros aplicados (para tipo 'filtrado')
        
    Returns:
        str: Nombre del archivo
    """
    fecha_actual = datetime.now().strftime("%Y%m%d_%H%M%S")
    
    if tipo == "completo":
        return f"analisis_validados_completo_{fecha_actual}.xlsx"
    
    elif tipo == "usuario" and usuario_correo:
        usuario_limpio = usuario_correo.split('@')[0]
        return f"analisis_validados_usuario_{usuario_limpio}_{fecha_actual}.xlsx"
    
    elif tipo == "filtrado":
        return f"analisis_validados_filtrado_{fecha_actual}.xlsx"
    
    else:
        return f"analisis_validados_{fecha_actual}.xlsx"


def validar_datos_para_excel(db: Session) -> Dict[str, Any]:
    """
    Valida que existan datos en la tabla de validados antes de generar Excel.
    
    Args:
        db (Session): Sesión de base de datos
        
    Returns:
        Dict: Información sobre la disponibilidad de datos
    """
    try:
        from src.AnalisisQuimicosValidados.infrastructure.analisis_quimicos_validados_model import (
            AnalisisQuimicosValidados,
        )
        
        # Contar total de registros
        total_validados = db.query(AnalisisQuimicosValidados).count()
        
        if total_validados == 0:
            return {
                "tiene_datos": False,
                "total_registros": 0,
                "mensaje": "No existen análisis validados en la base de datos"
            }
        
        # Obtener fecha del análisis más reciente
        analisis_reciente = (
            db.query(AnalisisQuimicosValidados)
            .order_by(AnalisisQuimicosValidados.fecha_validacion.desc())
            .first()
        )
        
        # Obtener fecha del análisis más antiguo
        analisis_antiguo = (
            db.query(AnalisisQuimicosValidados)
            .order_by(AnalisisQuimicosValidados.fecha_validacion.asc())
            .first()
        )
        
        return {
            "tiene_datos": True,
            "total_registros": total_validados,
            "fecha_mas_reciente": analisis_reciente.fecha_validacion.isoformat() if analisis_reciente and analisis_reciente.fecha_validacion else None,
            "fecha_mas_antigua": analisis_antiguo.fecha_validacion.isoformat() if analisis_antiguo and analisis_antiguo.fecha_validacion else None,
            "mensaje": f"Hay {total_validados} análisis validados disponibles para exportar"
        }
        
    except Exception as e:
        print(f"Error validando datos para Excel: {e}")
        return {
            "tiene_datos": False,
            "total_registros": 0,
            "error": str(e),
            "mensaje": "Error al verificar datos disponibles"
        }